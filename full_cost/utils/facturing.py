import importlib
import os
import datetime
from openpyxl import load_workbook, utils
from django.db.models import Q
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils.cell import get_column_letter
from io import BytesIO
import numpy as np
from datetime import date
from django.http import HttpResponse
from django.db.models import Max
from django.urls import reverse
from full_cost.constants.entities import Entity
# from full_cost.utils.constants import get_activities_from_entity, get_subbillings_from_entity_short,\
#     get_subbillings_from_entity_long, get_entity_long, CNRS_PERCENTAGE
from lab.models import Extraction, Price
from full_cost import settings

def get_border(style=None, color='FF000000'):
    return Border(left=Side(border_style=style, color=color),
           right=Side(border_style=style, color=color),
           top=Side(border_style=style, color=color),
           bottom=Side(border_style=style, color=color),
           diagonal=Side(border_style=style, color=color),
           diagonal_direction=0,
           outline=Side(border_style=style, color=color),
           vertical=Side(border_style=style, color=color),
           horizontal=Side(border_style=style, color=color)
           )

alignment = Alignment(horizontal='center',
                     vertical='center',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)


def as_text(value):
    return str(value) if value is not None else ""

def set_columns_width(worksheet):
    for column_cells in worksheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[utils.get_column_letter(column_cells[0].column)].width = length

def to_string(val):
    return '{:.02f}'.format(val)


def calculate_wus(records_list, entity: Entity):
    subbillings_long = get_subbillings_from_entity_long(entity)
    Nwu = np.array([0. for idx in range(len(subbillings_long))])
    for records in records_list:
        for r in records:
            if r.experiment.get_exp_type_display() in subbillings_long:
                ind = subbillings_long.index(r.experiment.get_exp_type_display())
                Nwutmp = np.array([r.wu if idx == ind else 0 for idx in range(len(subbillings_long))])
                Nwu += Nwutmp
    return Nwu

def populate_releve(records_list, project, entity, show_time=True):


    subbilling_long = get_subbillings_from_entity_long(entity)

    wb = load_workbook(filename=os.path.join(settings.STATIC_ROOT, 'template_facturation.xlsx'))
    ws = wb.create_sheet('Relevé')

    ws.append([None])
    ws.append([None])
    if entity == 'MECA' or entity == 'ELEC':
        header = ['Date', 'Worker', 'Session']
    else:
        header = ['Date', 'Experiment', 'Session']
    header.extend(subbilling_long)

    ws.append(header)
    Nwu = calculate_wus(records_list, entity)


    for records in records_list:
        records = records.order_by('experiment', 'date_from')
        for r in records:
            date_to = None
            date_from = r.date_from.strftime('%d/%m/%Y')
            time_to = None
            time_from = None
            if hasattr(r, 'time_from'):
                if isinstance(r.time_from, datetime.time):
                    time_from = r.time_from.strftime('%H:%M:%S')
                else:
                    time_from = r.get_time_from_display()

            if hasattr(r, 'date_to'):
                date_to = r.date_to.strftime('%d/%m/%Y')

            if hasattr(r, 'time_to'):
                if isinstance(r.time_to, datetime.time):
                    time_to = r.time_to.strftime('%H:%M:%S')
                else:
                    time_to = r.get_time_to_display()

            if date_to is not None:
                session = f"du {date_from}-{time_from if show_time else ''} au {date_to}-{time_to if show_time else ''}"
            else:
                session = f"le {date_from}: {time_from if show_time else ''} - {time_to if show_time else ''}"


            ind = subbilling_long.index(r.experiment.get_exp_type_display())
            wus = [r.wu if idx == ind else None for idx in range(len(subbilling_long))]
            row = [r.date_from, str(r.experiment.experiment), session]
            row.extend(wus)
            ws.append(row)

    res = [None, None, 'Total:']
    res.extend(Nwu)
    ws.append(res)
    last_cell = ws.calculate_dimension().split(':')[1]
    cell = ws[last_cell]
    letter = get_column_letter(cell.column)
    icol = cell.column
    irow = cell.row
    cells = ws['A3':f'{letter}{irow}']
    for row in cells:
        for cell in row:
            cell.border = get_border('medium')
            cell.alignment = alignment

    ws.append([None])
    set_columns_width(ws)
    ws['A1'] = f"Relevé des séances sur le projet: {str(project)}"
    ws['A1'].font = Font(name='Times New Roman', size=10, bold=False,)


    return wb, Nwu

def calculate_totals(project, records_list, entity):

    wus = calculate_wus(records_list, entity)
    subbilling_short = get_subbillings_from_entity_short(entity)
    totals = [0]

    for ind, bill in enumerate(subbilling_short):
        price, tarification = get_project_price(project, entity, bill)
        totals[0] += wus[ind]*price


    return totals

def get_project_price(project, entity, bill):
    if project.is_academic:
        tarification = 'académique'
        if project.is_national:
            tarification += ' nationale'
            price = Price.objects.get(price_entity=entity, price_category='T3ANR', price_name=bill).price
        else:
            tarification += ' internationale ou privée'
            price = Price.objects.get(price_entity=entity, price_category='T3', price_name=bill).price
        if not project.is_cnrs:
            price += price * CNRS_PERCENTAGE / 100
            tarification += ' non gérée par le CNRS'
        else:
            tarification += ' gérée par le CNRS'
    else:
        tarification = 'privée'
        price = Price.objects.get(price_entity=entity, price_category='T1', price_name=bill).price
    return price, tarification

def populate_facture(extraction_name, extraction, entity):
    records_list = []
    for act in get_activities_from_entity(entity):
        records_list.append(getattr(extraction, f'{act}_record_related').all())

    project = extraction.project

    dates = [extraction.date_after.strftime('%d/%m/%Y'),
            extraction.date_before.strftime('%d/%m/%Y'),]

    wb, wus = populate_releve(records_list, project, entity)


    ws = wb['Facture']

    totals = calculate_totals(project, records_list, entity)

    subbilling_short = get_subbillings_from_entity_short(entity)
    subbilling_long = get_subbillings_from_entity_long(entity)
    for ind, bill in enumerate(subbilling_short):
        price, tarification = get_project_price(project, entity, bill)
        row = [None, subbilling_long[ind], wus[ind], price, to_string(wus[ind]*price)]
        ws.append(row)


    ws.append([None, None, None, 'Total (€HT):', to_string(totals[0])])


    letter = 'E'
    irow = 24 + len(subbilling_long) + 1
    cells = ws['C24':f'{letter}{24}']
    for row in cells:
        for cell in row:
            cell.border = get_border('medium')
            cell.alignment = alignment

    cells = ws['B25':f'{letter}{irow}']
    for ind in range(25, irow+1):
        ws.row_dimensions[ind].height = 40
    for row in cells:
        for cell in row:
            cell.border = get_border('medium')
            cell.alignment = alignment

    ws['C13'] = str(project.project_pi)
    ws['C14'] = project.project_name
    ws['C17'] = get_entity_long(entity)
    ws['C20'] = dates[0]
    ws['E20'] = dates[1]
    ws['G6'] = date.today().strftime('%d/%m/%Y')
    ws['C2'] = extraction_name

    facture_object = '" et de "'.join(subbilling_long)
    ws['C19'] = f'Séances de "{facture_object}"'
    ws['B22'] = f'Tarification {tarification}'
    return wb


def export_book(wb):
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()

def generate_xlsx(extraction):
    ext_id = extraction.creation_id
    entity = extraction.billing
    extraction_name = f"{entity} {date.today().strftime('%y')}-{ext_id:03d}"

    wb = populate_facture(extraction_name, extraction, entity)
    data = export_book(wb)

    filename = f'extract_{extraction_name}.xlsx'
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment; filename="{}"'.format(filename)
    response.write(data)
    return response



def create_extraction(entity, records_list, project, filter):
    ext_id = Extraction.objects.all().filter(creation_date__year=date.today().year).aggregate(Max('creation_id'))['creation_id__max']
    if ext_id is not None:
        ext_id += 1
    else:
        ext_id = 0

    totals = calculate_totals(project, records_list, entity)
    total = totals[0]
    ext = Extraction(project=project,
                     date_after=filter.form.cleaned_data['date_from'].start,
                     date_before=filter.form.cleaned_data['date_from'].stop,
                     creation_id=ext_id, amount=total, billing=entity)

    ext.save()
    for records in records_list:
        for r in records:
            r.extraction = ext
            r.save()
    return ext